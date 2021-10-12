#!/usr/bin/env python3

#######################################################
#                                                     #
#             Presentation Grading Script             #
#                                                     #
#                    Written By:                      #
#              Caleb Kirby in Fall 2021               #
#                   github @ kirbk                    #
#                                                     #
#            Please review README.md for              #
#       possible required edits to this script.       #
#                                                     #
#######################################################


import openpyxl
from os import listdir
from os.path import isfile, join


MAX_POINTS = 100


# Look for .xlsx files in submission directory and keep track of them.
path = './submissions'
files = [f for f in listdir(path) if isfile(join(path, f)) and f.split('.')[-1] == 'xlsx']

workbooks = []

# Load excel documents into openpyxl.
for f in files:
  workbooks.append(openpyxl.load_workbook(filename = join(path, f), data_only=True))


invalid = False
invalid_sheets = {} # {sheet, book_idx}

book_idx = 0
students = {} # {"Name (Group#)", total_score}
for book in workbooks:
  perfect_count = 0 # This is the amount of perfect scores given by a single student.
  for sheet in book:
    i = 1
    while sheet.cell(row = 3, column = i).value != None:
      # c is the student's name, v is the score given by the student's sheet we are viewing.
      # !important I can see this being problematic in the future, if the row that the name
      # and total score is changed. But, this should be as simple as changing these values
      # accordingly.
      # Fair warning, this IS 1-indexed.

      c = sheet.cell(row = 3, column = i).value
      v = sheet.cell(row = 4, column = i).value

      # Several checks are gone through to determine if a sheet is valid.
      # If one inconsistency is found, the WHOLE sheet is thrown out of
      # the calculation. However, this is sometimes a little overzealous
      # so the log will show what book and what sheet the error happened
      # in. You should check these files and determine the fix, given the
      # information printed in the log. These errors are usually somebody
      # typing their own name somewhere that interferes with the name and
      # score finding part of this algorithm.

      # Checks to make sure the top cell is the name and that the lower cell
      # is numerical. Pretty sure there is a Python function to check if it's
      # numerical, instead of doing it this way, but this is what I wrote lol.
      if not (isinstance(c, str) and (isinstance(v, float) or isinstance(v, int))):
        invalid = True
        invalid_sheets[sheet] = book_idx
        print("Type mismatch in sheet:", sheet.title, "in book:", files[book_idx],
          "\n\t", type(c), "and", type(v))
        break
      
      # Originally, this checked to see if a score was *exactly* -14, which
      # is what it would be if they had not touched the sheet at all. 
      # However, I changed it to just check to see if the total score is
      # negative, as any negative value would likely be invalidating.
      # If you disagree, change this back.
      #
      # This is usually not an error, though, this is the expected result
      # when the student belongs to this group, as such, the score should
      # be ignored either way.
      if v < 10:
        invalid = True
        invalid_sheets[sheet] = book_idx
        print("Default value found in book:", files[book_idx],
          "\n\tin sheet", sheet.title, "\n\tvalue:", v)
        break
      
      # This check is very similar to the above one, except that it checks
      # for a value, other than the default value, in a student's book for
      # their own group. This is problematic, since the student should not
      # be evaluating themselves.
      if files[book_idx].find(c.strip().lower()) != -1:
        invalid = True
        invalid_sheets[sheet] = book_idx
        print("Wrongful non-default value found in book:", files[book_idx],
          "\n\tin sheet:", sheet.title, "\n\tvalue:", v)
        break

      # Checks for perfect scores, this is not an error, but too many for
      # one student could suggest lazy grading. This should be manually
      # kept track of, use your intuition. If you want, however, you can
      # add some warning code that tells you when it is above a certain
      # value. Right now, it just tells you the total for every student.
      if v == MAX_POINTS:
        perfect_count += 1

      # Scores can't be greater than the max score, this will thrown them
      # out and ignore them.
      if v > MAX_POINTS:
        invalid = True
        invalid_sheets[sheet] = book_idx
        print("Better-than-perfect score detected in book:", files[book_idx],
          "\n\tin sheet:", sheet.title, "\n\tscore:", v)
        break

      # Add the student's stripped name and group to a dictionary containing
      # their total score. If they are already in it, add to it.
      try:
        name = c.strip() + " (" + sheet.title + ")"
        students[name] = students.get(name, 0) + float(v)
      except Exception as e:
        # This will likely not run as the errors should be taken care of above,
        # but if it does, it will print out the except and what caused it.
        invalid = True
        invalid_sheets[sheet] = book_idx
        print("Error in book:", files[book_idx], "\n\tin sheet:", sheet.title, "\n\tvalue:", v)
        print("\t", e)
        break
  
      i += 1

  # Show how many perfect scores a student gave, for you records.
  print("\nBook", files[book_idx], "\n\thas", perfect_count, "perfect (or better) scores.")
  book_idx += 1


print()

# Prints out all of the sheets with bad information, these should
# be manually confirmed if you aren't lazy. They are likely
# simple fixes.
if invalid:
  print("Some sheets had bad information and were therefore ignored:")
  for bad in invalid_sheets:
    print("Bad sheet is in book:", files[invalid_sheets[bad]])
    print("\tBad sheet is:", bad.title)
    print()

# Prepare log files
l   = open("project.log", "w")
csv = open("project.csv", "w")

csv.write("Student,Total,Average,Denominator\n")

incorrect = []

# Attempt to reassign points that were erroneously divided.
for s in students:
  if students[s] < 300: # Usually, less than 300 means something went wrong.
    incorrect.append(s + ": " + str(students[s]))
    for cor in students:
      if cor != s:
        if cor.find(s[:-10]) != -1:
          # Look for matching student name, sans group, and apply the points to theirs.
          # This will also make it clear where the points went, so that you can, again,
          # manually confirm it. 
          incorrect[-1] = incorrect[-1] + " -> " + cor
          students[cor] = students.get(cor, 0) + float(students[s])
          students[s] = -1
          break

for s in students:
  if students[s] == -1:
    continue

  # Denominator for calcuations, this is how many sheets should affect your grades
  # and is based on the number of bad sheets.
  den = len(workbooks)

  # Decrement denominator if student belongs to group of invalidated sheet.
  for bad in invalid_sheets:
    if s.find(bad.title) != -1:
      den -= 1
      break

  
  # Basic log stuff.
  print(s + ":")
  print("\tTotal Score:", students[s])
  print("\tAverage Score:", round(students[s] / den, 3))
  print("\tDenominator:", den)

  l.write(s + ":\n")
  l.write("\tTotal Score: " + str(students[s]) + "\n")
  l.write("\tAverage Score: " + str(round(students[s] / den, 3)) + "\n")
  l.write("\tDenominator: " + str(den) + "\n")

  csv.write(s + "," + str(students[s]) + "," + str(students[s] / den) + "," + str(den) + "\n")

# Report the scores that were likely incorrect, most of these should have been
# automatically corrected, and will be denoted with an arrow (->) pointing
# to where these points ended up. If it was not corrected, you should confirm
# the score manually. It is probably the case the student was added in to a
# group later, or did not present.
if len(incorrect) > 0:
  print("\nSome reports were likely incorrect, they have been added to the predicted student:")

  for s in incorrect:
    print(s)


for book in workbooks:
  book.close()
